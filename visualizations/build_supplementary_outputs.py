from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = Path(__file__).with_name("outputs")
DATA_DIR = OUTPUT_DIR / "data"

ANALYSIS_XLSX = ROOT / "data" / "results" / "filtered_dataset_binary_classification_analysis.xlsx"
ARTICLE_XLSX = ROOT / "data" / "results" / "filtered_dataset_binary_classification.xlsx"
AI_FAMILY_MAP = ROOT / "sources" / "ai_family_map.csv"
THRESHOLDS_CSV = ROOT / "sources" / "thresholds.csv"
SOURCES_DIR = ROOT / "sources"
QUERY_PATH = ROOT / "documentation" / "self-documented docs" / "Web of Science search query.txt"
INSIGHTS_TXT = ROOT / "documentation" / "article" / "supplementary_results_insights.txt"

YEAR_COL = "Publication Year"
YEARS = list(range(2019, 2026))
EARLY_YEARS = [2019, 2020, 2021]
LATE_YEARS = [2023, 2024, 2025]
PERFORMANCE_EARLY_YEARS = [2019, 2020, 2021]
PERFORMANCE_LATE_YEARS = [2024, 2025]

SCORE_MAP = {
    "very low": 1,
    "low": 2,
    "medium": 3,
    "high": 4,
    "very high": 5,
    "Very Low": 1,
    "Low": 2,
    "Medium": 3,
    "High": 4,
    "Very High": 5,
}

SOURCE_PURPOSES = {
    "ai_family_map.csv": "Maps individual AI model-family flags to broader AI classes.",
    "ai_keywords.csv": "Curated model-family keyword dictionary used for article-level AI tagging.",
    "task_keywords.csv": "Curated keyword dictionary for primary and secondary task detection.",
    "cancer_keywords.csv": "Hard cancer-site dictionary used for cancer-site tagging.",
    "cancer_keywords_soft.csv": "Soft cancer-site dictionary for broader article-level cancer-site evidence.",
    "metric_synonyms.csv": "Metric synonym dictionary used by abstract-level performance-metric extraction.",
    "thresholds.csv": "Author-defined metric-specific ordinal proxy thresholds used to map raw abstract-level values to standardized reported-performance categories.",
    "task_metric_priority.csv": "Task-specific metric hierarchy used for composite and weighted categories.",
    "task_priority.csv": "Priority order used to select a single primary task when multiple tasks are detected.",
    "category_scores.csv": "Numeric scores assigned to ordinal performance categories.",
    "country_synonyms.csv": "Country-name harmonization overrides for corresponding-author geography.",
    "wos_exclusion_categories.tsv": "WoS categories used to support exclusion or ambiguity decisions.",
    "onco_terms_filter.csv": "Original oncology filtering term dictionary.",
    "onco_terms_filter_strong.csv": "Strong oncology evidence terms for eligibility filtering.",
    "onco_terms_filter_moderate.csv": "Moderate oncology evidence terms for eligibility filtering.",
    "onco_terms_filter_weak.csv": "Weak oncology evidence terms for eligibility filtering.",
    "onco_terms_filter_remove.csv": "Negative oncology terms used to reduce false inclusions.",
    "ai_terms_filter_strong.csv": "Strong AI evidence terms for eligibility filtering.",
    "ai_terms_filter_moderate.csv": "Moderate AI evidence terms for eligibility filtering.",
    "ai_terms_filter_weak.csv": "Weak AI evidence terms for eligibility filtering.",
    "ai_terms_filter_remove.csv": "Negative AI terms used to reduce false inclusions.",
    "raw_ai_terms_filter.csv": "Raw AI filtering term source retained for auditability.",
    "total-population-by-country-2025.csv": "2025 population denominator source used for per-capita country normalization.",
    "total-population-by-country-2025.csv": "2025 population denominator source used for per-capita country normalization.",
}


def ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    INSIGHTS_TXT.parent.mkdir(parents=True, exist_ok=True)


def save_csv(df: pd.DataFrame, name: str) -> Path:
    path = DATA_DIR / name
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def save_json(data: dict, name: str) -> Path:
    path = DATA_DIR / name
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def clean_label(value: object) -> str:
    return str(value).replace("_", " ")


def pct(value: float) -> str:
    return f"{value:.1f}%"


def load_sheet_columns(sheet_name: str) -> list[str]:
    return list(pd.read_excel(ANALYSIS_XLSX, sheet_name=sheet_name, nrows=0).columns)


def build_sheet_index() -> pd.DataFrame:
    wb = load_workbook(ANALYSIS_XLSX, read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows.append(
            {
                "Sheet": ws.title,
                "Rows excluding header": max(ws.max_row - 1, 0),
                "Columns": ws.max_column,
                "First columns": "; ".join(str(h) for h in headers[:8] if h is not None),
                "Suggested supplementary role": classify_sheet(ws.title),
            }
        )
    wb.close()
    return pd.DataFrame(rows)


def classify_sheet(sheet: str) -> str:
    lower = sheet.lower()
    if "country" in lower or "reprint" in lower:
        return "Corresponding-author geography"
    if "metric" in lower or "roc" in lower or "accuracy" in lower or "no metrics" in lower or "composite" in lower or "weighted" in lower:
        return "Metric reporting and reported performance"
    if "ai" in lower and "year" in lower:
        return "AI model/class temporal dynamics"
    if "cancer" in lower and "year" in lower:
        return "Cancer-site temporal dynamics"
    if "task" in lower:
        return "Task, cancer, and AI cross-tabulation"
    if "frequency" in lower or "distribution" in lower or "breakdown" in lower or "number of" in lower:
        return "Descriptor frequency table"
    if lower.startswith("meta"):
        return "Audit metadata frequency table"
    if "source title" in lower or "source" in lower:
        return "Journal/source-title summary"
    return "General aggregate output"


def build_dictionary_manifest() -> pd.DataFrame:
    rows = []
    for path in sorted(SOURCES_DIR.glob("*")):
        if not path.is_file() or path.name == ".gitkeep":
            continue
        if path.suffix.lower() not in {".csv", ".tsv"}:
            continue
        sep = "\t" if path.suffix.lower() == ".tsv" else None
        try:
            df = pd.read_csv(path, sep=sep, engine="python")
            rows.append(
                {
                    "File": str(path.relative_to(ROOT)),
                    "Rows excluding header": len(df),
                    "Columns": len(df.columns),
                    "Column names": "; ".join(map(str, df.columns[:12])),
                    "Purpose": SOURCE_PURPOSES.get(path.name, "Curated source file used by the reproducible pipeline."),
                }
            )
        except Exception as exc:  # pragma: no cover - manifest should keep going on malformed audit files.
            lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
            header = lines[0] if lines else ""
            delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
            columns = header.split(delimiter) if header else []
            rows.append(
                {
                    "File": str(path.relative_to(ROOT)),
                    "Rows excluding header": max(len(lines) - 1, 0),
                    "Columns": len(columns) if columns else "",
                    "Column names": "; ".join(columns),
                    "Purpose": SOURCE_PURPOSES.get(path.name, "Curated source file used by the reproducible pipeline.")
                    + " This file is treated as a line-oriented dictionary for manifest counting.",
                }
            )
    return pd.DataFrame(rows)


def _format_number(value: object) -> str:
    if pd.isna(value):
        return ""
    if value == np.inf or str(value).lower() == "inf":
        return "inf"
    numeric = float(value)
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:g}"


def build_metric_threshold_table() -> pd.DataFrame:
    thresholds = pd.read_csv(THRESHOLDS_CSV)
    label_order = ["Very High", "High", "Medium", "Low", "Very Low"]
    rows = []
    for metric, group in thresholds.groupby("metric", sort=False):
        group = group.copy()
        comparison = group["comparison"].iloc[0]
        direction = "Higher values indicate better reported performance" if comparison == "ge" else "Lower values indicate better reported performance"
        cutoffs = {row["label"]: row["cutoff"] for _, row in group.iterrows()}
        ranges: dict[str, str] = {}
        if comparison == "ge":
            for i, label in enumerate(label_order):
                cutoff = cutoffs[label]
                if label == "Very High":
                    ranges[label] = f">= {_format_number(cutoff)}"
                elif label == "Very Low":
                    previous = cutoffs["Low"]
                    ranges[label] = f"< {_format_number(previous)}"
                else:
                    previous = cutoffs[label_order[i - 1]]
                    ranges[label] = f">= {_format_number(cutoff)} and < {_format_number(previous)}"
        else:
            for i, label in enumerate(label_order):
                cutoff = cutoffs[label]
                if label == "Very High":
                    ranges[label] = f"<= {_format_number(cutoff)}"
                elif label == "Very Low":
                    previous = cutoffs["Low"]
                    ranges[label] = f"> {_format_number(previous)}"
                else:
                    previous = cutoffs[label_order[i - 1]]
                    ranges[label] = f"> {_format_number(previous)} and <= {_format_number(cutoff)}"
        rows.append(
            {
                "Metric": metric,
                "Direction": direction,
                "Very high": ranges["Very High"],
                "High": ranges["High"],
                "Medium": ranges["Medium"],
                "Low": ranges["Low"],
                "Very low": ranges["Very Low"],
                "Source": "Author-defined ordinal proxy threshold encoded in sources/thresholds.csv",
            }
        )
    return pd.DataFrame(rows)


def derive_inputs() -> tuple[pd.DataFrame, list[str], list[str], list[str]]:
    ai_models = [c for c in load_sheet_columns("AI Models by Year") if c != YEAR_COL]
    cancers = [c for c in load_sheet_columns("Task x Cancer") if c != "Task Category"]
    family_map = pd.read_csv(AI_FAMILY_MAP)
    class_to_models = {
        ai_class: [m for m in family_map.loc[family_map["main_family"] == ai_class, "subfamily_column"].tolist() if m in ai_models]
        for ai_class in family_map["main_family"].dropna().unique()
    }

    base_cols = [
        YEAR_COL,
        "weighted_category",
        "composite_metric",
        "composite_source",
        "no_metrics_reported",
        "primary_task",
    ]
    requested = base_cols + ai_models + cancers
    seen = set()
    requested = [c for c in requested if not (c in seen or seen.add(c))]
    df = pd.read_excel(ARTICLE_XLSX, usecols=requested)

    class_df = {}
    for ai_class, models in class_to_models.items():
        class_df[ai_class] = (df[models].sum(axis=1) > 0).astype(int) if models else 0
    df = pd.concat([df, pd.DataFrame(class_df, index=df.index)], axis=1)
    df["_weighted_score"] = df["weighted_category"].map(SCORE_MAP)
    return df, ai_models, list(class_to_models.keys()), cancers


def build_trend_table(df: pd.DataFrame, items: list[str], label: str) -> pd.DataFrame:
    year_counts = df[YEAR_COL].value_counts().sort_index().reindex(YEARS, fill_value=0)
    rows = []
    grouped = df.groupby(YEAR_COL)
    for col in items:
        counts = grouped[col].sum().reindex(YEARS, fill_value=0).astype(int)
        total = int(counts.sum())
        if total == 0:
            continue
        shares = counts / year_counts.replace(0, np.nan) * 100
        nonzero = counts[counts > 0]
        early = int(counts.loc[EARLY_YEARS].sum())
        late = int(counts.loc[LATE_YEARS].sum())
        late_early_ratio = np.inf if early == 0 else (late / len(LATE_YEARS)) / (early / len(EARLY_YEARS))
        rows.append(
            {
                label: col,
                "Total count": total,
                "Share of full corpus (%)": round(total / len(df) * 100, 2),
                "2019 count": int(counts.get(2019, 0)),
                "2025 count": int(counts.get(2025, 0)),
                "2019 share (%)": round(float(shares.get(2019, 0)), 2),
                "2025 share (%)": round(float(shares.get(2025, 0)), 2),
                "2019 to 2025 share change (pp)": round(float(shares.get(2025, 0) - shares.get(2019, 0)), 2),
                "First detected year": int(nonzero.index.min()),
                "Peak year": int(counts.idxmax()),
                "Peak count": int(counts.max()),
                "2019-2021 count": early,
                "2023-2025 count": late,
                "Late/early annualized ratio": round(float(late_early_ratio), 2) if np.isfinite(late_early_ratio) else "Inf",
            }
        )
    return pd.DataFrame(rows)


def build_performance_table(df: pd.DataFrame, items: list[str], label: str, min_total: int = 20) -> pd.DataFrame:
    rows = []
    for col in items:
        sub = df[df[col] == 1]
        total = len(sub)
        if total < min_total:
            continue
        scoreable = sub[sub["_weighted_score"].notna()]
        vc = scoreable["weighted_category"].value_counts()
        scoreable_n = len(scoreable)
        high_vhigh = int(vc.get("very high", 0) + vc.get("high", 0))
        rows.append(
            {
                label: col,
                "Total labelled records": total,
                "Scoreable records": scoreable_n,
                "Scoreable share of labelled records (%)": round(scoreable_n / total * 100, 1) if total else 0,
                "Very high": int(vc.get("very high", 0)),
                "High": int(vc.get("high", 0)),
                "Medium": int(vc.get("medium", 0)),
                "Low": int(vc.get("low", 0)),
                "Very low": int(vc.get("very low", 0)),
                "High or very high share of scoreable records (%)": round(high_vhigh / scoreable_n * 100, 1) if scoreable_n else np.nan,
                "Very high share of scoreable records (%)": round(int(vc.get("very high", 0)) / scoreable_n * 100, 1) if scoreable_n else np.nan,
                "Mean weighted category score": round(float(scoreable["_weighted_score"].mean()), 2) if scoreable_n else np.nan,
            }
        )
    return pd.DataFrame(rows)


def build_performance_change_table(df: pd.DataFrame, model_trends: pd.DataFrame) -> pd.DataFrame:
    rows = []
    top_models = model_trends.sort_values("Total count", ascending=False).head(25)["AI Model"].tolist()
    for model in top_models:
        sub = df[df[model] == 1]
        row = {"AI Model": model}
        for label, years in [("Early 2019-2021", PERFORMANCE_EARLY_YEARS), ("Late 2024-2025", PERFORMANCE_LATE_YEARS)]:
            period = sub[sub[YEAR_COL].isin(years)]
            scoreable = period[period["_weighted_score"].notna()]
            row[f"{label} total"] = len(period)
            row[f"{label} scoreable"] = len(scoreable)
            row[f"{label} mean score"] = round(float(scoreable["_weighted_score"].mean()), 2) if len(scoreable) else np.nan
            row[f"{label} high/very high (%)"] = round(float(scoreable["weighted_category"].isin(["high", "very high"]).mean() * 100), 1) if len(scoreable) else np.nan
        row["Late minus early mean score"] = round(row["Late 2024-2025 mean score"] - row["Early 2019-2021 mean score"], 2)
        row["Late minus early high/very high (pp)"] = round(row["Late 2024-2025 high/very high (%)"] - row["Early 2019-2021 high/very high (%)"], 1)
        rows.append(row)
    return pd.DataFrame(rows)


def top_records(df: pd.DataFrame, sort_col: str, n: int = 6, ascending: bool = False) -> list[dict]:
    return df.sort_values(sort_col, ascending=ascending).head(n).to_dict(orient="records")


def build_key_insights(
    model_trends: pd.DataFrame,
    class_trends: pd.DataFrame,
    cancer_trends: pd.DataFrame,
    model_perf: pd.DataFrame,
    class_perf: pd.DataFrame,
    cancer_perf: pd.DataFrame,
    perf_change: pd.DataFrame,
) -> dict:
    min_model = model_trends["Total count"] >= 20
    min_class = class_trends["Total count"] >= 20
    min_cancer = cancer_trends["Total count"] >= 50
    min_model_perf = model_perf["Total labelled records"] >= 100
    min_class_perf = class_perf["Total labelled records"] >= 100
    min_cancer_perf = cancer_perf["Total labelled records"] >= 100
    return {
        "top_ai_models_overall": top_records(model_trends, "Total count"),
        "fastest_rising_ai_models_by_share_change": top_records(model_trends[min_model], "2019 to 2025 share change (pp)"),
        "largest_relative_declines_ai_models": top_records(
            model_trends[(model_trends["Total count"] >= 100) & (model_trends["2019 count"] > 0)],
            "2019 to 2025 share change (pp)",
            ascending=True,
        ),
        "fastest_rising_ai_classes_by_share_change": top_records(class_trends[min_class], "2019 to 2025 share change (pp)"),
        "largest_relative_declines_ai_classes": top_records(
            class_trends[(class_trends["Total count"] >= 100) & (class_trends["2019 count"] > 0)],
            "2019 to 2025 share change (pp)",
            ascending=True,
        ),
        "fastest_rising_cancer_sites_by_share_change": top_records(cancer_trends[min_cancer], "2019 to 2025 share change (pp)"),
        "largest_relative_declines_cancer_sites": top_records(
            cancer_trends[(cancer_trends["Total count"] >= 50) & (cancer_trends["2019 count"] > 0)],
            "2019 to 2025 share change (pp)",
            ascending=True,
        ),
        "highest_reported_weighted_performance_ai_models_min100": top_records(
            model_perf[min_model_perf], "High or very high share of scoreable records (%)"
        ),
        "lowest_reported_weighted_performance_ai_models_min100": top_records(
            model_perf[min_model_perf],
            "High or very high share of scoreable records (%)",
            ascending=True,
        ),
        "highest_reported_weighted_performance_ai_classes_min100": top_records(
            class_perf[min_class_perf], "High or very high share of scoreable records (%)"
        ),
        "lowest_reported_weighted_performance_ai_classes_min100": top_records(
            class_perf[min_class_perf],
            "High or very high share of scoreable records (%)",
            ascending=True,
        ),
        "highest_reported_weighted_performance_cancers_min100": top_records(
            cancer_perf[min_cancer_perf], "High or very high share of scoreable records (%)"
        ),
        "lowest_reported_weighted_performance_cancers_min100": top_records(
            cancer_perf[min_cancer_perf],
            "High or very high share of scoreable records (%)",
            ascending=True,
        ),
        "largest_late_vs_early_performance_gains_top_models": top_records(perf_change, "Late minus early mean score"),
    }


def write_insight_text(insights: dict) -> None:
    lines = [
        "Supplementary results insights",
        "==============================",
        "",
        "Interpretation note",
        "- Performance categories are abstract-level reported metric categories, not independent model benchmarking, risk-of-bias assessment, or clinical effectiveness.",
        "- Cancer-site and AI-family labels are multi-label descriptors, so counts are not mutually exclusive.",
        "",
        "AI model dynamics",
    ]

    for row in insights["fastest_rising_ai_models_by_share_change"][:8]:
        lines.append(
            f"- {clean_label(row['AI Model'])}: {row['2019 count']} records in 2019 ({row['2019 share (%)']}%) and "
            f"{row['2025 count']} in 2025 ({row['2025 share (%)']}%); change {row['2019 to 2025 share change (pp)']} pp."
        )
    lines.extend(["", "AI class dynamics"])
    for row in insights["fastest_rising_ai_classes_by_share_change"][:6]:
        lines.append(
            f"- {clean_label(row['AI Class'])}: {row['2019 share (%)']}% in 2019 to {row['2025 share (%)']}% in 2025 "
            f"({row['2019 to 2025 share change (pp)']} pp)."
        )
    lines.extend(["", "Cancer-site dynamics"])
    for row in insights["fastest_rising_cancer_sites_by_share_change"][:8]:
        lines.append(
            f"- {row['Cancer Type']}: {row['2019 share (%)']}% in 2019 to {row['2025 share (%)']}% in 2025 "
            f"({row['2019 to 2025 share change (pp)']} pp)."
        )
    lines.extend(["", "Reported weighted performance by AI model, minimum 100 labelled records"])
    for row in insights["highest_reported_weighted_performance_ai_models_min100"][:8]:
        lines.append(
            f"- {clean_label(row['AI Model'])}: {row['High or very high share of scoreable records (%)']}% high/very high "
            f"among {row['Scoreable records']} scoreable records; scoreable share {row['Scoreable share of labelled records (%)']}%."
        )
    lines.extend(["", "Lower reported weighted performance by AI model, minimum 100 labelled records"])
    for row in insights["lowest_reported_weighted_performance_ai_models_min100"][:8]:
        lines.append(
            f"- {clean_label(row['AI Model'])}: {row['High or very high share of scoreable records (%)']}% high/very high "
            f"among {row['Scoreable records']} scoreable records; scoreable share {row['Scoreable share of labelled records (%)']}%."
        )
    INSIGHTS_TXT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    ensure_dirs()
    sheet_index = build_sheet_index()
    dictionary_manifest = build_dictionary_manifest()
    df, ai_models, ai_classes, cancers = derive_inputs()

    model_trends = build_trend_table(df, ai_models, "AI Model")
    class_trends = build_trend_table(df, ai_classes, "AI Class")
    cancer_trends = build_trend_table(df, cancers, "Cancer Type")
    model_perf = build_performance_table(df, ai_models, "AI Model")
    class_perf = build_performance_table(df, ai_classes, "AI Class")
    cancer_perf = build_performance_table(df, cancers, "Cancer Type")
    perf_change = build_performance_change_table(df, model_trends)

    save_csv(sheet_index, "supplementary_analysis_workbook_sheet_index.csv")
    save_csv(dictionary_manifest, "supplementary_dictionary_manifest.csv")
    save_csv(build_metric_threshold_table(), "supplementary_metric_thresholds_readable.csv")
    save_csv(model_trends.sort_values("Total count", ascending=False), "supplementary_ai_model_trends.csv")
    save_csv(class_trends.sort_values("Total count", ascending=False), "supplementary_ai_class_trends.csv")
    save_csv(cancer_trends.sort_values("Total count", ascending=False), "supplementary_cancer_site_trends.csv")
    save_csv(model_perf.sort_values("Total labelled records", ascending=False), "supplementary_weighted_performance_by_ai_model.csv")
    save_csv(class_perf.sort_values("Total labelled records", ascending=False), "supplementary_weighted_performance_by_ai_class.csv")
    save_csv(cancer_perf.sort_values("Total labelled records", ascending=False), "supplementary_weighted_performance_by_cancer_site.csv")
    save_csv(perf_change, "supplementary_ai_model_performance_early_vs_late.csv")

    insights = build_key_insights(model_trends, class_trends, cancer_trends, model_perf, class_perf, cancer_perf, perf_change)
    save_json(
        {
            "corpus_n": int(len(df)),
            "years": YEARS,
            "web_of_science_query_file": str(QUERY_PATH.relative_to(ROOT)),
            "insights": insights,
        },
        "supplementary_key_insights.json",
    )
    write_insight_text(insights)
    print(f"Wrote supplementary outputs to {DATA_DIR}")
    print(f"Wrote insight notes to {INSIGHTS_TXT}")


if __name__ == "__main__":
    main()
